# inventory_manager.py

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import os

inventory = {}  # {serial_number: {"product": product_name, "quantity": qty}}
sales_history = []  # Track sales with serial numbers

EXCEL_FILE = "sales_data.xlsx"

# ---------- FUNCTIONS ----------

def add_product():
    serial_number = serial_entry.get().strip()
    product_name = product_entry.get().strip()
    qty = quantity_entry.get().strip()
    price_value = price_entry.get().strip()

    if not serial_number or not product_name or not qty:
        messagebox.showerror("Input Error", "Please fill the required fields.")
        return

    if serial_number in inventory:
        messagebox.showerror("Error", "This serial number already exists!")
        return

    try:
        qty = int(qty)
        if qty <= 0:
            raise ValueError
    except:
        messagebox.showerror("Input Error", "Quantity must be a positive number.")
        return

    price = 0.0
    if price_value:
        try:
            price = float(price_value)
            if price < 0:
                raise ValueError
        except:
            messagebox.showerror("Input Error", "Price must be a non-negative number.")
            return

    inventory[serial_number] = {"product": product_name, "quantity": qty, "price": price}
    refresh_table()
    save_to_excel_auto()

    serial_entry.delete(0, tk.END)
    product_entry.delete(0, tk.END)
    quantity_entry.delete(0, tk.END)
    price_entry.delete(0, tk.END)


def process_sale():
    serial = serial_sale_entry.get().strip()
    customer = customer_entry.get().strip()
    quantity = sale_quantity_entry.get().strip()

    if not serial:
        messagebox.showwarning(
            "Serial Number Required",
            "Please enter serial number."
        )
        return

    if serial not in inventory:
        messagebox.showerror(
            "Error",
            f"Serial number '{serial}' not found in inventory."
        )
        return

    if not customer:
        messagebox.showwarning(
            "Customer Name Required",
            "Please enter customer name."
        )
        return

    if not quantity:
        messagebox.showwarning(
            "Quantity Required",
            "Please enter quantity to sell."
        )
        return

    try:
        quantity = int(quantity)
        if quantity <= 0:
            raise ValueError
    except:
        messagebox.showerror(
            "Input Error",
            "Quantity must be a positive number."
        )
        return

    # Check if enough quantity available
    available_qty = inventory[serial]["quantity"]
    if available_qty < quantity:
        messagebox.showerror(
            "Insufficient Stock",
            f"Only {available_qty} units available for serial {serial}. Cannot sell {quantity} units."
        )
        return

    product_name = inventory[serial]["product"]
    unit_price = inventory[serial].get("price", 0.0)
    total_price = quantity * unit_price
    
    # Reduce quantity
    inventory[serial]["quantity"] -= quantity
    refresh_table()

    time_now = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    # Record sale
    sale_record = {
        "serial": serial,
        "customer": customer,
        "product": product_name,
        "quantity": quantity,
        "price": unit_price,
        "total_price": total_price,
        "time": time_now
    }
    sales_history.append(sale_record)

    # Update treeview
    history.insert(
        "",
        tk.END,
        values=(serial, customer, product_name, quantity, time_now)
    )

    # Auto-save to Excel
    save_to_excel_auto()

    serial_sale_entry.delete(0, tk.END)
    customer_entry.delete(0, tk.END)
    sale_quantity_entry.delete(0, tk.END)

    messagebox.showinfo(
        "Sale Successful",
        f"Sale recorded for {customer}.\n{quantity} unit(s) of {product_name} (Serial: {serial}) sold."
    )

    if inventory[serial]["quantity"] == 0:
        messagebox.showwarning(
            "Stock Over",
            f"Stock is over for {product_name} (Serial: {serial})."
        )


def refresh_table():
    for row in table.get_children():
        table.delete(row)

    for serial, info in inventory.items():
        table.insert(
            "",
            tk.END,
            values=(serial, info["product"], info["quantity"])
        )


def open_edit_inventory_dialog(event=None):
    selected = table.selection()
    if not selected:
        return

    item_id = selected[0]
    values = table.item(item_id)["values"]
    if not values:
        return

    old_serial = str(values[0]).strip()
    old_product = str(values[1]).strip()
    old_qty = str(values[2]).strip()

    edit_window = tk.Toplevel(root)
    edit_window.title("Edit Inventory Item")
    edit_window.geometry("420x220")
    edit_window.resizable(False, False)
    edit_window.transient(root)
    edit_window.grab_set()

    tk.Label(edit_window, text="Serial Number").grid(row=0, column=0, padx=10, pady=10, sticky="e")
    serial_edit = tk.Entry(edit_window, width=28)
    serial_edit.grid(row=0, column=1, pady=10)
    serial_edit.insert(0, old_serial)

    tk.Label(edit_window, text="Product Name").grid(row=1, column=0, padx=10, pady=10, sticky="e")
    product_edit = tk.Entry(edit_window, width=28)
    product_edit.grid(row=1, column=1, pady=10)
    product_edit.insert(0, old_product)

    tk.Label(edit_window, text="Quantity").grid(row=2, column=0, padx=10, pady=10, sticky="e")
    quantity_edit = tk.Entry(edit_window, width=28)
    quantity_edit.grid(row=2, column=1, pady=10)
    quantity_edit.insert(0, old_qty)

    def save_edit():
        new_serial = serial_edit.get().strip()
        new_product = product_edit.get().strip()
        new_quantity = quantity_edit.get().strip()

        if not new_serial:
            new_serial = old_serial
        else:
            new_serial = str(new_serial)

        if not new_product or not new_quantity:
            messagebox.showerror("Input Error", "Product name and quantity are required.")
            return

        try:
            new_quantity = int(new_quantity)
            if new_quantity < 0:
                raise ValueError
        except:
            messagebox.showerror("Input Error", "Quantity must be a non-negative integer.")
            return

        if new_serial != old_serial and new_serial in inventory:
            messagebox.showerror("Error", "Serial number already exists.")
            return

        item_info = inventory.pop(old_serial)
        item_info["product"] = new_product
        item_info["quantity"] = new_quantity
        inventory[new_serial] = item_info

        for sale in sales_history:
            if sale["serial"] == old_serial:
                sale["serial"] = new_serial
                sale["product"] = new_product

        refresh_table()
        history.delete(*history.get_children())
        for sale in sales_history:
            history.insert(
                "",
                tk.END,
                values=(sale["serial"], sale["customer"], sale["product"], sale["quantity"], sale["time"])
            )
        save_to_excel_auto()
        edit_window.destroy()

    save_button = tk.Button(edit_window, text="Save Changes", command=save_edit, bg="#2563eb", fg="white", width=18)
    save_button.grid(row=3, column=0, columnspan=2, pady=15)


def search_inventory():
    query = search_entry.get().strip().lower()
    if not query:
        messagebox.showwarning("Search Required", "Please enter a serial number or product name to search.")
        return

    found = False
    for item in table.get_children():
        values = table.item(item)["values"]
        if not values:
            continue
        serial = str(values[0]).lower()
        product = str(values[1]).lower()
        if query in serial or query in product:
            table.selection_set(item)
            table.focus(item)
            table.see(item)
            found = True
            break

    if not found:
        messagebox.showinfo("Not Found", "No inventory item matches the search query.")


def show_home():
    messagebox.showinfo(
        "Home",
        "Welcome to Stock Management!\n\nUse Add Product or Import Excel to load inventory.\nSearch by Serial or Product Name.\nDouble-click an item to edit it."
    )


def import_excel_file(filepath=None):
    if not filepath:
        filepath = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
        )

    if not filepath:
        return

    try:
        wb = load_workbook(filepath, data_only=True)
        if "Inventory" not in wb.sheetnames:
            messagebox.showerror("Import Error", "The selected file must contain an 'Inventory' sheet.")
            return

        inventory.clear()
        sales_history.clear()
        history.delete(*history.get_children())

        inv_sheet = wb["Inventory"]
        headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(inv_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        serial_index = headers.index("serial number") if "serial number" in headers else 0
        product_index = headers.index("product name") if "product name" in headers else 1 if len(headers) > 1 else 1
        qty_index = headers.index("quantity") if "quantity" in headers else 2 if len(headers) > 2 else 2
        price_index = headers.index("price") if "price" in headers else None

        for row in inv_sheet.iter_rows(min_row=2, values_only=True):
            if not row or row[serial_index] is None:
                continue
            serial = str(row[serial_index]).strip()
            product_name = str(row[product_index]).strip() if row[product_index] is not None else ""
            quantity = int(row[qty_index]) if row[qty_index] is not None else 0
            price = 0.0
            if price_index is not None and row[price_index] is not None:
                try:
                    price = float(row[price_index])
                except:
                    price = 0.0
            inventory[serial] = {"product": product_name, "quantity": quantity, "price": price}

        if "Sales History" in wb.sheetnames:
            sales_sheet = wb["Sales History"]
            sales_headers = [str(cell.value).strip().lower() if cell.value is not None else "" for cell in next(sales_sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
            s_serial_index = sales_headers.index("serial number") if "serial number" in sales_headers else 0
            s_customer_index = sales_headers.index("customer name") if "customer name" in sales_headers else 1
            s_product_index = sales_headers.index("product") if "product" in sales_headers else 2
            s_qty_index = sales_headers.index("quantity") if "quantity" in sales_headers else 3
            s_price_index = sales_headers.index("price") if "price" in sales_headers else None
            s_time_index = sales_headers.index("date & time") if "date & time" in sales_headers else 4

            for row in sales_sheet.iter_rows(min_row=2, values_only=True):
                if not row or row[s_serial_index] is None:
                    continue
                serial = str(row[s_serial_index]).strip()
                customer = str(row[s_customer_index]).strip() if row[s_customer_index] is not None else ""
                product_name = str(row[s_product_index]).strip() if row[s_product_index] is not None else ""
                quantity = int(row[s_qty_index]) if row[s_qty_index] is not None else 0
                price = 0.0
                if s_price_index is not None and row[s_price_index] is not None:
                    try:
                        price = float(row[s_price_index])
                    except:
                        price = 0.0
                time_value = str(row[s_time_index]) if row[s_time_index] is not None else ""
                sale_record = {
                    "serial": serial,
                    "customer": customer,
                    "product": product_name,
                    "quantity": quantity,
                    "price": price,
                    "time": time_value
                }
                sales_history.append(sale_record)
                history.insert(
                    "",
                    tk.END,
                    values=(serial, customer, product_name, quantity, time_value)
                )

        refresh_table()
        save_to_excel_auto()
        messagebox.showinfo("Import Successful", "Inventory imported successfully from Excel.")
    except Exception as e:
        messagebox.showerror("Import Error", f"Failed to import Excel: {str(e)}")


def show_import_instructions():
    guide = tk.Toplevel(root)
    guide.title("Import Excel Instructions")
    guide.geometry("520x420")
    guide.resizable(False, False)
    guide.transient(root)
    guide.grab_set()

    instruction_text = (
        "To import Excel into Stock Management:\n\n"
        "1. Use an .xlsx file.\n"
        "2. Add a sheet named 'Inventory'.\n"
        "3. Inventory sheet columns should include:\n"
        "   - Serial Number\n"
        "   - Product Name\n"
        "   - Quantity\n"
        "   - Price (optional, will be saved but not shown in app)\n\n"
        "4. You may also include a 'Sales History' sheet with:\n"
        "   - Serial Number\n"
        "   - Customer Name\n"
        "   - Product\n"
        "   - Quantity\n"
        "   - Price (optional)\n"
        "   - Total Price (optional)\n"
        "   - Date & Time\n\n"
    )

    text_widget = tk.Text(guide, wrap="word", padx=12, pady=12, bg="#f4f6f9", bd=0)
    text_widget.insert("1.0", instruction_text)
    text_widget.configure(state="disabled")
    text_widget.pack(fill="both", expand=True)

    button_frame = tk.Frame(guide, bg="#f4f6f9")
    button_frame.pack(fill="x", pady=10)

    tk.Button(button_frame, text="Import Excel File", command=lambda: import_excel_file(), bg="#2563eb", fg="white", width=18).pack(side="left", padx=10)
    tk.Button(button_frame, text="Close", command=guide.destroy, bg="#6b7280", fg="white", width=18).pack(side="right", padx=10)


def save_to_excel_auto():
    """Automatically save inventory and sales to Excel without prompting"""
    try:
        inv_headers = ["Serial Number", "Product Name", "Quantity", "Price"]
        sales_headers = ["Serial Number", "Customer Name", "Product", "Quantity", "Price", "Total Price", "Date & Time"]

        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            if "Inventory" not in wb.sheetnames:
                inventory_sheet = wb.create_sheet("Inventory")
                inventory_sheet.append(inv_headers)
            else:
                inventory_sheet = wb["Inventory"]
            if "Sales History" not in wb.sheetnames:
                sales_sheet = wb.create_sheet("Sales History")
                sales_sheet.append(sales_headers)
            else:
                sales_sheet = wb["Sales History"]

            if inventory_sheet.max_row == 1 and all(cell.value is None for cell in inventory_sheet[1]):
                inventory_sheet.append(inv_headers)
            if sales_sheet.max_row == 1 and all(cell.value is None for cell in sales_sheet[1]):
                sales_sheet.append(sales_headers)
        else:
            wb = Workbook()
            inventory_sheet = wb.active
            inventory_sheet.title = "Inventory"
            inventory_sheet.append(inv_headers)
            sales_sheet = wb.create_sheet("Sales History")
            sales_sheet.append(sales_headers)

        # Style headers
        header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in inventory_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for cell in sales_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Clear existing data rows in both sheets
        if inventory_sheet.max_row > 1:
            inventory_sheet.delete_rows(2, inventory_sheet.max_row - 1)
        if sales_sheet.max_row > 1:
            sales_sheet.delete_rows(2, sales_sheet.max_row - 1)

        # Add inventory data
        for serial, info in inventory.items():
            inventory_sheet.append([serial, info["product"], info["quantity"], info.get("price", 0.0)])

        # Add sales data
        for sale in sales_history:
            sales_sheet.append([
                sale["serial"],
                sale["customer"],
                sale["product"],
                sale["quantity"],
                sale.get("price", 0.0),
                sale.get("total_price", 0.0),
                sale["time"]
            ])

        # Adjust column widths
        inventory_sheet.column_dimensions["A"].width = 20
        inventory_sheet.column_dimensions["B"].width = 30
        inventory_sheet.column_dimensions["C"].width = 15
        inventory_sheet.column_dimensions["D"].width = 15
        sales_sheet.column_dimensions["A"].width = 20
        sales_sheet.column_dimensions["B"].width = 25
        sales_sheet.column_dimensions["C"].width = 30
        sales_sheet.column_dimensions["D"].width = 15
        sales_sheet.column_dimensions["E"].width = 15
        sales_sheet.column_dimensions["F"].width = 15
        sales_sheet.column_dimensions["G"].width = 30

        # Center align numeric columns in inventory and sales
        for row in inventory_sheet.iter_rows(min_row=2, max_row=inventory_sheet.max_row, min_col=3, max_col=4):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
        for row in sales_sheet.iter_rows(min_row=2, max_row=sales_sheet.max_row, min_col=4, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        wb.save(EXCEL_FILE)
    except Exception as e:
        messagebox.showerror(
            "Save Error",
            f"Failed to save to Excel: {str(e)}"
        )


def load_from_excel():
    """Load inventory and sales history from the auto-save Excel file."""
    if not os.path.exists(EXCEL_FILE):
        return

    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)

        if "Inventory" in wb.sheetnames:
            ws = wb["Inventory"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                serial = str(row[0]).strip()
                product_name = str(row[1]).strip() if row[1] is not None else ""
                qty = int(row[2]) if row[2] is not None else 0
                price = float(row[3]) if row[3] is not None else 0.0
                inventory[serial] = {"product": product_name, "quantity": qty, "price": price}

        if "Sales History" in wb.sheetnames:
            ws = wb["Sales History"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                serial = str(row[0]).strip()
                customer = str(row[1]).strip() if row[1] is not None else ""
                product_name = str(row[2]).strip() if row[2] is not None else ""
                quantity = int(row[3]) if row[3] is not None else 0
                price = float(row[4]) if row[4] is not None else 0.0
                total_price = float(row[5]) if len(row) > 5 and row[5] is not None else quantity * price
                time_index = 6 if len(row) > 6 else 5
                time_value = str(row[time_index]) if len(row) > time_index and row[time_index] is not None else ""
                sale_record = {
                    "serial": serial,
                    "customer": customer,
                    "product": product_name,
                    "quantity": quantity,
                    "price": price,
                    "total_price": total_price,
                    "time": time_value
                }
                sales_history.append(sale_record)
                history.insert(
                    "",
                    tk.END,
                    values=(serial, customer, product_name, quantity, time_value)
                )

        refresh_table()
    except Exception as e:
        messagebox.showerror(
            "Load Error",
            f"Failed to load inventory from Excel: {str(e)}"
        )





# ---------- WINDOW ----------

root = tk.Tk()
root.title("Stock Management")
root.geometry("1100x650")
root.configure(bg="#f4f6f9")

menubar = tk.Menu(root)
menubar.add_command(label="Home", command=show_home)
menubar.add_command(label="Import Excel", command=import_excel_file)
menubar.add_command(label="Import Instructions", command=show_import_instructions)
root.config(menu=menubar)

# ---------- HEADER ----------

header = tk.Label(
    root,
    text="Stock Management",
    font=("Segoe UI", 22, "bold"),
    bg="#1f2937",
    fg="white",
    pady=15
)

header.pack(fill="x")

# ---------- ADD PRODUCT ----------

add_frame = tk.LabelFrame(
    root,
    text="Add Product",
    padx=15,
    pady=15,
    font=("Segoe UI", 10, "bold"),
    bg="#f4f6f9"
)

add_frame.pack(fill="x", padx=20, pady=15)

tk.Label(
    add_frame,
    text="Serial Number",
    bg="#f4f6f9"
).grid(row=0, column=0, padx=10)

serial_entry = tk.Entry(add_frame, width=20)
serial_entry.grid(row=0, column=1)

tk.Label(
    add_frame,
    text="Product Name",
    bg="#f4f6f9"
).grid(row=0, column=2, padx=10)

product_entry = tk.Entry(add_frame, width=25)
product_entry.grid(row=0, column=3)

tk.Label(
    add_frame,
    text="Quantity",
    bg="#f4f6f9"
).grid(row=0, column=4, padx=10)

quantity_entry = tk.Entry(add_frame, width=10)
quantity_entry.grid(row=0, column=5)

tk.Label(
    add_frame,
    text="Price",
    bg="#f4f6f9"
).grid(row=0, column=6, padx=10)

price_entry = tk.Entry(add_frame, width=12)
price_entry.grid(row=0, column=7)

add_button = tk.Button(
    add_frame,
    text="Add Product",
    command=add_product,
    bg="#2563eb",
    fg="white",
    width=15
)

add_button.grid(row=0, column=8, padx=20)

# ---------- INVENTORY TABLE ----------

table_frame = tk.LabelFrame(
    root,
    text="Current Inventory",
    padx=10,
    pady=10,
    font=("Segoe UI", 10, "bold"),
    bg="#f4f6f9"
)

table_frame.pack(fill="both", padx=20, pady=10)

search_frame = tk.Frame(table_frame, bg="#f4f6f9")
search_frame.pack(fill="x", pady=(0, 10))

tk.Label(
    search_frame,
    text="Search Inventory",
    bg="#f4f6f9"
).pack(side="left", padx=(0, 10))

search_entry = tk.Entry(search_frame, width=30)
search_entry.pack(side="left")
search_entry.bind("<Return>", lambda event: search_inventory())

search_button = tk.Button(
    search_frame,
    text="Find",
    command=search_inventory,
    bg="#2563eb",
    fg="white",
    width=12
)
search_button.pack(side="left", padx=10)

clear_button = tk.Button(
    search_frame,
    text="Clear",
    command=lambda: search_entry.delete(0, tk.END),
    bg="#6b7280",
    fg="white",
    width=12
)
clear_button.pack(side="left")

table = ttk.Treeview(
    table_frame,
    columns=("Serial", "Product", "Qty"),
    show="headings",
    height=8
)

table.heading("Serial", text="Serial Number")
table.heading("Product", text="Product Name")
table.heading("Qty", text="Quantity")

table.column("Serial", width=150)
table.column("Product", width=300)
table.column("Qty", width=150)

table.pack(fill="x")

table.bind("<Double-1>", open_edit_inventory_dialog)

# ---------- SALES SECTION ----------

sales_frame = tk.LabelFrame(
    root,
    text="Process Sale",
    padx=15,
    pady=15,
    font=("Segoe UI", 10, "bold"),
    bg="#f4f6f9"
)

sales_frame.pack(fill="x", padx=20, pady=10)

tk.Label(
    sales_frame,
    text="Serial Number",
    bg="#f4f6f9"
).grid(row=0, column=0, padx=10)

serial_sale_entry = tk.Entry(sales_frame, width=20)
serial_sale_entry.grid(row=0, column=1)

tk.Label(
    sales_frame,
    text="Customer Name",
    bg="#f4f6f9"
).grid(row=0, column=2, padx=10)

customer_entry = tk.Entry(sales_frame, width=25)
customer_entry.grid(row=0, column=3)

tk.Label(
    sales_frame,
    text="Quantity",
    bg="#f4f6f9"
).grid(row=0, column=4, padx=10)

sale_quantity_entry = tk.Entry(sales_frame, width=15)
sale_quantity_entry.grid(row=0, column=5)

sale_button = tk.Button(
    sales_frame,
    text="Complete Sale",
    command=process_sale,
    bg="#16a34a",
    fg="white",
    width=15
)

sale_button.grid(row=0, column=6, padx=10)

import_button = tk.Button(
    sales_frame,
    text="Import Excel",
    command=import_excel_file,
    bg="#ea580c",
    fg="white",
    width=15
)

import_button.grid(row=0, column=7, padx=5)

# ---------- SALES HISTORY ----------

history_frame = tk.LabelFrame(
    root,
    text="Sales History",
    padx=10,
    pady=10,
    font=("Segoe UI", 10, "bold"),
    bg="#f4f6f9"
)

history_frame.pack(fill="both", expand=True, padx=20, pady=10)

history = ttk.Treeview(
    history_frame,
    columns=("Serial", "Customer", "Product", "Qty", "Date"),
    show="headings",
    height=10
)

history.heading("Serial", text="Serial Number")
history.heading("Customer", text="Customer Name")
history.heading("Product", text="Product")
history.heading("Qty", text="Quantity")
history.heading("Date", text="Date & Time")

history.column("Serial", width=120)
history.column("Customer", width=180)
history.column("Product", width=200)
history.column("Qty", width=80)
history.column("Date", width=220)

history.pack(fill="both", expand=True)

# Load inventory and sales automatically if data exists
load_from_excel()

# ---------- START ----------

root.mainloop()